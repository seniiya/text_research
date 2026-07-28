"""코딩용 시트를 만든다 (CSV + XLSX).

문장 분리는 split_sentences.py(코딩북 3항)를 그대로 쓴다.
이 스크립트가 더하는 것은 '문맥 행'뿐이다.

배치 — 각 turn 맨 위에 그 turn의 사용자 발화를 한 줄 넣는다:

    turn  sent_no  role      text                       code
    1     -        Prompt    요즘 잘 움직이지도 않고…       -      <- 문맥. 코딩 대상 아님
    1     1        Response  그렇게 느껴질 만해.                  <- 코딩 대상
    1     2        Response  특히 요즘 활동량이 많이…

사용자 발화는 코딩하지 않지만(코딩북 §2) 화면에는 보여야 한다.
RST(재진술)·RFL(감정 반영)·QST(되묻기)는 정의 자체가 사용자 발화를
기준으로 하므로, 문맥 없이는 판정할 수 없다.

XLSX 쪽에는 코딩 편의 장치를 넣는다:
  - code 열 드롭다운(8개 코드) -> 오타 불가
  - 문맥 행 음영 + 시트 보호 -> 코딩 대상 셀에만 입력 가능
  - 헤더 고정, text 열 줄바꿈
  - '코드표' 시트에 코딩북 §4~§6 요약

사용: python scripts/make_coding_sheet.py P01
      (XLSX 생성에는 openpyxl 필요. 없으면 CSV만 만든다.)
"""

import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from split_sentences import split_response  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
CHATS = ROOT / "data" / "deid" / "chats"
OUT_DIR = ROOT / "data" / "coding"

# esconv : ESConv 8전략 라벨. 드롭다운으로 고른다.
# note   : 판정이 애매했던 이유를 적는 자유기술. 비워도 된다.
#          체계가 데이터에 맞는지 진단할 때(3항 7항) 근거 자료가 된다.
# 둘 다 Response 행에만 입력한다. 문맥 행은 '-' 로 잠근다.
FIELDS = ["p_id", "turn", "sent_no", "role", "text", "chars", "esconv", "note"]
LABEL_FIELDS = ["esconv", "note"]
DROPDOWN_FIELDS = ["esconv"]        # note 는 자유기술이라 목록을 걸지 않는다

CODES = [
    ("QST", "질문", "사용자에게 정보나 감정을 되묻는 발화", "탐색"),
    ("RST", "재진술", "사용자가 한 말을 다시 정리해 되돌려 줌", "탐색"),
    ("RFL", "감정 반영", "사용자의 감정을 짚어 말해 줌", "정서적 지원"),
    ("SLF", "자기개방", "화자 자신의 경험·감정을 공유", "정서적 지원"),
    ("AFF", "인정·안심", "사용자를 긍정하거나 걱정을 덜어 줌", "정서적 지원"),
    ("SUG", "제안 제시", "구체적 행동·해결책 제안", "도구적 지원"),
    ("INF", "정보 제공", "사실·원인·수치 등 정보 전달", "도구적 지원"),
    ("OTH", "기타", "인사, 농담, 이모지, 목록 제목 등", "기타"),
]

PRIORITY = [
    "물음표로 끝나면 → QST",
    "사용자 말을 되돌려 정리하면 → RST",
    "감정어(힘들겠다, 이해돼, 답답하지)가 중심이면 → RFL",
    "걱정을 덜어 주거나 긍정하면 → AFF",
    "'~해봐', '~하는 게 좋아' → SUG",
    "사실·원인·수치 설명 → INF",
    "위에 해당 없음 → OTH",
]


def flatten(text):
    """문맥 행은 한 셀에 들어가므로 줄바꿈을 공백으로 합친다."""
    return " ".join(line.strip() for line in text.split("\n") if line.strip())


def build_rows(pid):
    doc = json.loads((CHATS / f"{pid}.json").read_text(encoding="utf-8"))
    is_md = doc.get("metadata", {}).get("format") == "markdown"
    rows, turn, pending = [], 0, None

    for msg in doc["messages"]:
        if msg["role"] == "Prompt":
            pending = msg["say"]
            continue
        if msg["role"] != "Response":
            continue

        turn += 1
        if pending is not None:
            ctx = flatten(pending)
            rows.append({
                "p_id": pid, "turn": turn, "sent_no": "-", "role": "Prompt",
                "text": ctx, "chars": len(ctx), "esconv": "-", "note": "-",
            })
            pending = None

        for n, s in enumerate(split_response(msg["say"], markdown=is_md), start=1):
            rows.append({
                "p_id": pid, "turn": turn, "sent_no": n, "role": "Response",
                "text": s, "chars": len(s), "esconv": "", "note": "",
            })
    return rows


def carry_over(pid, rows):
    """이미 입력된 라벨을 새 시트로 옮긴다.

    문장 분리 규칙이 바뀌면 행이 다시 만들어진다. 그때 입력해 둔 라벨이
    사라지지 않도록, 기존 XLSX 를 읽어 (turn, sent_no, text) 가 같은 행에
    값을 되돌려 놓는다. text 까지 비교하므로 분리 결과가 달라진 행은
    옮기지 않는다. 잘못된 위치에 라벨이 붙는 것보다 비는 편이 낫다.
    """
    path = OUT_DIR / f"{pid}.xlsx"
    if not path.exists():
        return 0, 0

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return 0, 0

    ws = load_workbook(path)["coding"]
    header = [c.value for c in ws[1]]
    old = {}
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=i).value
               for i, h in enumerate(header, start=1) if h}
        if row.get("role") != "Response":
            continue
        vals = {f: row.get(f) for f in LABEL_FIELDS if row.get(f)}
        if vals:
            old[(str(row.get("turn")), str(row.get("sent_no")), row.get("text"))] = vals

    kept = 0
    for r in rows:
        if r["role"] != "Response":
            continue
        vals = old.pop((str(r["turn"]), str(r["sent_no"]), r["text"]), None)
        if vals:
            r.update(vals)
            kept += 1
    return kept, len(old)


def write_csv(pid, rows):
    path = OUT_DIR / f"{pid}.csv"
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    return path


def write_xlsx(pid, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "coding"

    head_fill = PatternFill("solid", fgColor="2F4858")
    ctx_fill = PatternFill("solid", fgColor="E8E4DC")
    band_fill = PatternFill("solid", fgColor="F7F7F5")
    code_fill = PatternFill("solid", fgColor="FFF9E6")

    ws.append(list(FIELDS))  # CSV와 열 이름을 똑같이 유지한다
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([r[f] for f in FIELDS])

    widths = {"p_id": 7, "turn": 6, "sent_no": 8, "role": 10,
              "text": 72, "chars": 7, "esconv": 10, "note": 30}
    for i, f in enumerate(FIELDS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[f]

    text_col = FIELDS.index("text") + 1
    label_cols = [FIELDS.index(f) + 1 for f in LABEL_FIELDS]

    for i, r in enumerate(rows, start=2):
        is_ctx = r["role"] == "Prompt"
        for j in range(1, len(FIELDS) + 1):
            cell = ws.cell(row=i, column=j)
            cell.alignment = Alignment(
                wrap_text=(j == text_col),
                vertical="top",
                horizontal="left" if j == text_col else "center",
            )
            cell.font = Font(size=10, italic=is_ctx,
                             color="6B6255" if is_ctx else "000000")
            if is_ctx:
                cell.fill = ctx_fill
            elif r["turn"] % 2 == 0:
                cell.fill = band_fill

        for col in label_cols:
            cell = ws.cell(row=i, column=col)
            if is_ctx:
                cell.protection = Protection(locked=True)
            else:
                # 코딩 대상 셀만 잠금 해제한다. 시트 보호가 켜져 있으므로
                # 코더는 이 셀들에만 입력할 수 있다.
                cell.protection = Protection(locked=False)
                cell.fill = code_fill

    last = len(rows) + 1
    for field, col in zip(LABEL_FIELDS, label_cols):
        if field not in DROPDOWN_FIELDS:
            continue
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(c[0] for c in CODES) + '"',
            allow_blank=True,
            showDropDown=False,   # False = 드롭다운 화살표를 보여준다 (openpyxl 규약)
            errorTitle="코드 아님",
            error="ESConv 8개 코드 중에서 고르세요.",
            promptTitle=field,
            prompt="QST RST RFL SLF AFF SUG INF OTH",
        )
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{last}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{len(rows) + 1}"
    # 비밀번호는 걸지 않는다. 규칙을 바꿔야 하면 [검토]-[시트 보호 해제]로 풀 수 있다.
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    # ── 참고 시트 ──
    ref = wb.create_sheet("코드표")
    ref.append(["코드", "전략", "정의", "상위 축"])
    for c in ref[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in CODES:
        ref.append(list(row))
    for i, w in enumerate([8, 12, 46, 14], start=1):
        ref.column_dimensions[get_column_letter(i)].width = w
    for row in ref.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10)

    ref.append([])
    ref.append(["판정 우선순위 (위에서부터 먼저 걸리는 규칙을 따른다)"])
    ref.cell(row=ref.max_row, column=1).font = Font(bold=True, size=10)
    for n, line in enumerate(PRIORITY, start=1):
        ref.append([n, line])
    ref.append([])
    ref.append(["esconv", "위 8개 코드 중 하나를 고른다."])
    ref.cell(row=ref.max_row, column=1).font = Font(bold=True, size=10)
    ref.append([])
    ref.append(["문맥 행(회색, role=Prompt)은 사용자 발화이며 코딩 대상이 아니다."])
    ref.append(["코딩 대상은 role=Response 행뿐이다. (코딩북 §2)"])

    path = OUT_DIR / f"{pid}.xlsx"
    wb.save(path)
    return path


def main():
    pid = sys.argv[1] if len(sys.argv) > 1 else "P01"
    rows = build_rows(pid)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    n_ctx = sum(1 for r in rows if r["role"] == "Prompt")
    n_code = sum(1 for r in rows if r["role"] == "Response")

    kept, lost = carry_over(pid, rows)
    csv_path = write_csv(pid, rows)
    xlsx_path = write_xlsx(pid, rows)

    print(f"{pid}: 총 {len(rows)}행 = 문맥 {n_ctx}행 + 코딩 대상 {n_code}행")
    if kept or lost:
        print(f"  기존 라벨 {kept}건 이어받음"
              + (f" / {lost}건은 해당 문장이 바뀌어 옮기지 못함" if lost else ""))
    print(f"  CSV : {csv_path.relative_to(ROOT).as_posix()}")
    if xlsx_path:
        print(f"  XLSX: {xlsx_path.relative_to(ROOT).as_posix()}  (드롭다운·시트보호 적용)")
    else:
        print("  XLSX: 건너뜀 (openpyxl 없음)")


if __name__ == "__main__":
    main()
