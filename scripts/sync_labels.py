"""XLSX 에 입력된 라벨을 같은 이름의 CSV 에 옮긴다.

코더는 엑셀에서 작업한다. CSV 는 시트를 만든 시점에 멈춰 있어서, 코딩이
진행될수록 XLSX 와 벌어진다. 실제로 coder1/P01 은 CSV 44건 / XLSX 183건까지
벌어져 있었다.

시트를 다시 만들지 않는다. 키로 행을 찾아 라벨 열만 채운다. 문장이 다르면 그
행은 건너뛴다. 잘못된 위치에 라벨이 붙는 것보다 비는 편이 낫다.

두 종류의 시트를 다룬다. 파일명으로 구분한다.

    P##.xlsx            제1부 — AI 응답 전략 (esconv). 키 (turn, sent_no)
    P##_speechact.xlsx  제4부 — 사용자 화행·감정 표면성. 키 utterance_id

사용: python scripts/sync_labels.py P01 [--coder=coder2]
      python scripts/sync_labels.py P04_speechact --coder=coder1
      python scripts/sync_labels.py --all
"""

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CODING = ROOT / "data" / "coding"

SPEECHACT_SUFFIX = "_speechact"


def spec_for(stem, coder):
    """시트 종류별 설정. key 는 행을 찾는 열, labels 는 옮길 열."""
    if stem.endswith(SPEECHACT_SUFFIX):
        return {
            "key": ["utterance_id"],
            "labels": [f"speech_act_{coder}", f"emotion_surface_{coder}", "note"],
            "count": f"speech_act_{coder}",
            "only_role": None,
        }
    return {
        "key": ["turn", "sent_no"],
        "labels": ["esconv", "note"],
        "count": "esconv",
        # 제1부의 코딩 대상은 Response 행뿐이다 (코딩북 §2).
        "only_role": "Response",
    }


def read_xlsx(path, spec):
    from openpyxl import load_workbook

    ws = load_workbook(path)["coding"]
    head = [c.value for c in ws[1]]

    need = list(spec["key"]) + ["text"]
    if spec["only_role"]:
        need.append("role")
    missing = [c for c in need if c not in head]
    if missing:
        raise KeyError(", ".join(missing))

    col = {k: head.index(k) for k in need}
    # note 열은 나중에 생겼다. 그 전에 만든 시트에는 없다.
    col.update({f: head.index(f) for f in spec["labels"] if f in head})

    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if spec["only_role"] and r[col["role"]] != spec["only_role"]:
            continue
        key = tuple(str(r[col[k]]) for k in spec["key"])
        out[key] = {
            "text": r[col["text"]] or "",
            **{f: (str(r[col[f]]).strip() if r[col[f]] else "")
               for f in spec["labels"] if f in col},
        }
    return out


def sync(stem, coder):
    xp, cp = CODING / coder / f"{stem}.xlsx", CODING / coder / f"{stem}.csv"
    if not xp.exists() or not cp.exists():
        return None

    spec = spec_for(stem, coder)
    try:
        src = read_xlsx(xp, spec)
    except KeyError as e:
        print(f"[{coder}] {stem}: 건너뜀 — 예상한 열이 없다 ({e})")
        return None

    with open(cp, encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        fields, rows = reader.fieldnames, list(reader)

    moved = changed = 0
    for r in rows:
        if spec["only_role"] and r.get("role") != spec["only_role"]:
            continue
        v = src.get(tuple(str(r[k]) for k in spec["key"]))
        if not v:
            continue
        if v["text"] != r["text"]:
            moved += 1          # 문장이 다르다. 붙이지 않는다.
            continue
        for f in spec["labels"]:
            if f in v and v[f] != (r.get(f) or ""):
                r[f] = v[f]
                changed += 1

    if changed:
        with open(cp, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=fields)
            w.writeheader()
            w.writerows(rows)

    labeled = sum(1 for r in rows
                  if not (spec["only_role"] and r.get("role") != spec["only_role"])
                  and (r.get(spec["count"]) or "").strip())
    return changed, moved, labeled


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a.split("=")[0]: a.split("=", 1)[-1]
             for a in sys.argv[1:] if a.startswith("--")}

    if "--all" in flags:
        targets = [(p.stem, p.parent.name)
                   for p in sorted(CODING.glob("*/*.xlsx"))]
    else:
        targets = [(args[0] if args else "P01", flags.get("--coder", "coder1"))]

    total = 0
    for stem, coder in targets:
        res = sync(stem, coder)
        if res is None:
            continue
        changed, moved, labeled = res
        if changed or moved:
            print(f"[{coder}] {stem}: {changed}칸 갱신, 라벨 {labeled}건"
                  + (f" / {moved}행은 문장이 달라 건너뜀" if moved else ""))
            total += changed
    print(f"총 {total}칸 갱신." if total else "갱신할 것 없음. CSV 가 XLSX 와 일치한다.")


if __name__ == "__main__":
    main()
