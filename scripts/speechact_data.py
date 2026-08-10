"""분석 스크립트가 공유하는 것 — 데이터 로더와 결과 기록 장치.

**로더가 여기 있는 이유.** 분석 스크립트가 여럿이라 로딩을 각자 구현하면 제외
규칙이 어긋난다. '규칙 6 붙여넣은 문서는 빈칸' 같은 규칙은 한 군데서만 틀려도
문장 수가 달라져 보고서 수치가 갈린다. 그래서 읽기는 전부 이 파일을 거친다.
정본은 XLSX 다. CSV 는 sync_labels.py 가 뽑은 사본이라 여기서는 XLSX 만 읽는다.

**Results 가 여기 있는 이유.** 수치를 표준출력으로만 내면 발표·논문에 옮겨 적는
순간 원본과 끊긴다(1차 보고서에서 실제로 문장 3개가 어긋났다). 그래서 분석
스크립트는 주장할 만한 값을 Results 에 적어 조각 파일로 남기고,
collect_results.py 가 그걸 모아 results.csv · RESULTS.md 를 만든다.
"""

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CODER1 = ROOT / "data" / "coding" / "coder1"
SURVEY_CSV = ROOT / "data" / "deid" / "survey_analysis.csv"
# 출력은 세 곳뿐이다. 늘리지 않는다.
#   tables/  사람이 열어 보는 표(csv). 파일 이름 앞머리가 어느 분석인지 말한다.
#   report/  발표·보고용 산출물(docx · png).
#   _build/  스크립트끼리 주고받는 중간 파일. 직접 열 일이 없다.
TABLES = ROOT / "data" / "output" / "tables"
REPORT_DIR = ROOT / "data" / "output" / "report"
BUILD = ROOT / "data" / "output" / "_build"
OUT_DIR = TABLES

# 축 A. 목록 순서가 곧 판정 우선순위이자 표·차트의 정렬 순서다.
SPEECH_ACT = ["FRAG", "RESP", "PHATIC", "EXPR", "Q", "C", "RQ", "RC", "S", "UNC"]
CODE_MEANING = {
    "FRAG": "절이 완결되지 않은 발화",
    "RESP": "직전 AI 발화에 대한 대답·수락·거절",
    "PHATIC": "인사·작별·감사·사과",
    "EXPR": "감정·평가의 분출이 발화의 목적",
    "Q": "실제로 답을 요구",
    "C": "행동·수행을 요구",
    "RQ": "의문문 형식이나 답을 요구하지 않음",
    "RC": "명령문 형식이나 실제 지시가 아님",
    "S": "정보·상태 서술",
    "UNC": "위 어디에도 안 맞음",
}
# 3i4K 원 체계에 있는 6코드. 원 코퍼스와 대조할 때 이것만 쓴다.
CODES_3I4K = ["FRAG", "S", "Q", "C", "RQ", "RC"]
# 대화행위 표준에서 가져와 채운 3코드 (ISO 24617-2 · DAMSL · Searle 1976).
CODES_ADDED = ["RESP", "PHATIC", "EXPR"]
SURFACE = ["EXP", "IMP", "NONE"]

# 3i4K 원 코퍼스 분포 (Cho et al. 2018, IU 제외 후 재정규화). 논문 수치이므로 상수다.
CORPUS_3I4K = {"FRAG": 2.2, "S": 45.3, "Q": 20.1, "C": 25.8, "RQ": 3.5, "RC": 3.2}
CORPUS_3I4K_N = 17735


def _cell(v):
    return v.strip() if isinstance(v, str) else v


def sheet_rows(path):
    """코딩 시트 한 장을 dict 목록으로. 빈 라벨은 None 이 아니라 '' 로 통일한다.

    prev_assistant_turn 은 시트에서 **턴의 첫 문장에만** 적혀 있다(같은 값을 문장마다
    반복하면 읽을 수 없기 때문). 분석에서 't08-s02 의 직전 AI 응답 길이'를 물으면
    0 이 나와 버리므로, 여기서 턴 안으로 채워 넣는다.
    """
    ws = load_workbook(path, data_only=True)["coding"]
    head = [c.value for c in ws[1]]
    out = []
    ctx, turn = "", None
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {h: _cell(v) for h, v in zip(head, r)}
        if not d.get("text"):
            continue
        if d.get("turn_index") != turn:
            turn, ctx = d.get("turn_index"), d.get("prev_assistant_turn") or ""
        d["prev_assistant_turn"] = d.get("prev_assistant_turn") or ctx
        out.append(d)
    return out


def load_labels(include_unlabeled=False):
    """{p_id: [행, ...]}.

    기본값은 **화행이 채워진 행만** 준다. 규칙 6(붙여넣은 문서)에 걸린 행은
    두 축이 모두 빈칸이라 여기서 자동으로 빠진다.
    include_unlabeled=True 면 미코딩 참가자를 포함한 전체를 준다 (진행률 계산용).
    """
    out = {}
    for x in sorted(CODER1.glob("P*_speechact.xlsx")):
        if "_v1" in x.name:                       # 원 체계 대조본은 별도 함수로
            continue
        pid = x.name.split("_")[0]
        rows = sheet_rows(x)
        for d in rows:
            d["p_id"] = pid
        out[pid] = rows if include_unlabeled else [
            d for d in rows if d.get("speech_act_coder1")]
    return {k: v for k, v in out.items() if v or include_unlabeled}


def load_v1(pid="P04"):
    """원 체계(v1) 코딩본. P04 한 명뿐이다."""
    p = CODER1 / f"{pid}_speechact_v1.xlsx"
    return sheet_rows(p) if p.exists() else []


def load_survey():
    with open(SURVEY_CSV, encoding="utf-8-sig") as fh:
        return {r["p_id"]: r for r in csv.DictReader(fh)}


def num(s):
    try:
        return float(str(s).strip())
    except (ValueError, AttributeError, TypeError):
        return None


# ── 결과 기록 ─────────────────────────────────────────────────────────
RESULTS_DIR = BUILD / "results"

# RESULTS.md 의 절 순서. 분석 스크립트는 어느 절에 넣을지만 정하고,
# 순서와 배치는 collect_results.py 가 이 목록대로 한다.
SECTIONS = [
    "참가자",
    "평소 AI와 무슨 대화를 하는가",
    "이번에 실제로 한 대화",
    "감정 자기보고 VA",
    "사용자 발화 화행",
    "정서 표면성과 자기보고",
]


def fmt_num(v, nd=1):
    """r·V 처럼 1 미만인 값을 소수 첫째로 반올림하면 0.668 이 0.7 이 돼 의미가
    사라진다. 자릿수를 값의 크기로 정한다."""
    if isinstance(v, bool) or not isinstance(v, float):
        return str(v)
    if v == 0:
        return "0"
    return f"{v:.3f}" if abs(v) < 1 else f"{v:.{nd}f}"


class Results:
    """분석 스크립트가 RESULTS.md 에 실을 것을 적어 두는 곳.

        R = Results("analyze_survey.py")
        R.table("참가자", "인구통계", ["구분", "값"], [["성별", "F 16 · M 6"]])
        R.num("사용자 발화 화행", "물음표인데 Q 가 아닌 비율", 23.1, "%", n=78)
        R.save("S")
    """

    def __init__(self, source):
        self.source, self.items = source, []

    def num(self, section, item, value, unit="", n=None, fig="", note=""):
        assert section in SECTIONS, f"모르는 절: {section}"
        self.items.append(dict(kind="num", section=section, item=item,
                               value=fmt_num(value), unit=unit,
                               n="" if n is None else n, fig=fig, note=note,
                               source=self.source))

    def table(self, section, title, headers, rows, note="", fig=""):
        """표는 부르는 쪽이 이미 round() 로 자릿수를 정해 넘긴다. 여기서 다시
        반올림하면 안 되므로 %g 로 표기만 다듬는다 — 5.0 은 '5', 0.9 는 '0.9'."""
        assert section in SECTIONS, f"모르는 절: {section}"
        def cell(c):
            if c is None:
                return ""
            if isinstance(c, float):
                return f"{c:g}"
            return str(c)
        self.items.append(dict(kind="table", section=section, title=title,
                               headers=list(headers),
                               rows=[[cell(c) for c in r] for r in rows],
                               note=note, fig=fig, source=self.source))

    def save(self, key):
        """key 는 RESULTS.md 안에서의 순서를 정한다 (1_설문 · 2_화행 · 3_표면성)."""
        RESULTS_DIR.mkdir(parents=True, exist_ok=True)
        p = RESULTS_DIR / f"{key}.json"
        with open(p, "w", encoding="utf-8") as fh:
            json.dump(self.items, fh, ensure_ascii=False, indent=1)
        nt = sum(1 for i in self.items if i["kind"] == "table")
        print(f"\n결과 기록: {p.relative_to(ROOT)}  "
              f"(표 {nt} · 수치 {len(self.items) - nt})")
        return p
