"""코더 간 일치도를 계산한다 (Cohen's kappa).

라벨은 XLSX 에서 읽는다. 코더는 엑셀에서 작업하므로 CSV 는 시트를 다시 만든
시점에 멈춰 있을 수 있다. 실제로 coder1/P01 은 CSV 44건 / XLSX 183건이었다.

두 코더가 같은 (turn, sent_no) 를 같은 문장에 대해 매겼는지 먼저 확인하고,
어긋나면 계산하지 않는다. 행이 밀린 채 나온 kappa 는 아무 의미가 없다.

사용: python scripts/agreement.py P01 [--a=coder1] [--b=coder2]
"""

import csv
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CODING = ROOT / "data" / "coding"
PRIVATE = ROOT / "data" / "private"

CODES = ["QST", "RST", "RFL", "SLF", "AFF", "SUG", "INF", "OTH"]

# Landis & Koch (1977). 논문에 kappa 를 쓸 때 관례적으로 붙는 구간.
BANDS = [
    (0.81, "almost perfect  거의 완전 일치"),
    (0.61, "substantial      상당함"),
    (0.41, "moderate         보통"),
    (0.21, "fair             약함"),
    (0.00, "slight           미미함"),
    (-1.0, "poor             우연 이하"),
]


def read_labels(path):
    """XLSX 의 coding 시트에서 Response 행만 읽는다."""
    from openpyxl import load_workbook

    ws = load_workbook(path)["coding"]
    head = [c.value for c in ws[1]]
    col = {k: head.index(k) for k in ("turn", "sent_no", "role", "text", "esconv")}

    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[col["role"]] != "Response":
            continue
        key = (str(r[col["turn"]]), str(r[col["sent_no"]]))
        out[key] = {
            "text": r[col["text"]] or "",
            "esconv": str(r[col["esconv"]]).strip() if r[col["esconv"]] else "",
        }
    return out


def editor(path):
    """엑셀이 남긴 마지막 저장자. 파일을 누가 만졌는지 확인용."""
    import re
    import zipfile

    try:
        body = zipfile.ZipFile(path).read("docProps/core.xml").decode("utf-8", "ignore")
    except Exception:
        return ""
    m = re.search(r"<cp:lastModifiedBy>(.*?)</cp:lastModifiedBy>", body)
    return m.group(1).strip() if m else ""


def self_coded(pid, paths):
    """코더가 그 참여자 본인인지 본다.

    본인은 그때 자기가 뭘 원했는지 알기 때문에, 텍스트만 보고 판정해야 하는
    RST·RFL·QST 를 정상적으로 코딩할 수 없다. 연결키가 있을 때만 확인한다
    (data/private/ 는 Git 에 없다).
    """
    keyfile = PRIVATE / "mapping.csv"
    if not keyfile.exists():
        return []
    with open(keyfile, encoding="utf-8", newline="") as fh:
        name = {r["p_id"]: r["name"] for r in csv.DictReader(fh)}
    who = name.get(pid, "")
    return [p for p in paths if who and editor(p) == who] if who else []


def kappa(pairs):
    """Cohen's kappa. pairs = [(코더A 라벨, 코더B 라벨), ...]"""
    n = len(pairs)
    po = sum(1 for a, b in pairs if a == b) / n
    ca, cb = Counter(a for a, _ in pairs), Counter(b for _, b in pairs)
    pe = sum((ca[c] / n) * (cb[c] / n) for c in set(ca) | set(cb))
    return po, pe, (po - pe) / (1 - pe) if pe != 1 else float("nan")


def band(k):
    for lo, label in BANDS:
        if k >= lo:
            return label
    return BANDS[-1][1]


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a.split("=")[0]: a.split("=", 1)[-1]
             for a in sys.argv[1:] if a.startswith("--")}

    pid = args[0] if args else "P01"
    a_name, b_name = flags.get("--a", "coder1"), flags.get("--b", "coder2")
    pa, pb = CODING / a_name / f"{pid}.xlsx", CODING / b_name / f"{pid}.xlsx"

    for p in (pa, pb):
        if not p.exists():
            sys.exit(f"[중단] 없음: {p.relative_to(ROOT).as_posix()}")

    A, B = read_labels(pa), read_labels(pb)

    if set(A) != set(B):
        sys.exit(f"[중단] 행 구성이 다르다. {a_name} {len(A)}행 / {b_name} {len(B)}행."
                 f" 같은 시트에서 코딩했는지 확인할 것.")

    moved = [k for k in A if A[k]["text"] != B[k]["text"]]
    if moved:
        sys.exit(f"[중단] 같은 (turn, sent_no) 인데 문장이 다른 행 {len(moved)}개."
                 f" 행이 밀렸다. 예: {moved[:3]}")

    both = sorted(k for k in A if A[k]["esconv"] and B[k]["esconv"])
    only_a = [k for k in A if A[k]["esconv"] and not B[k]["esconv"]]
    only_b = [k for k in A if B[k]["esconv"] and not A[k]["esconv"]]

    print(f"{pid}  {a_name} vs {b_name}")
    print(f"  코딩 대상 {len(A)}행 | 양쪽 라벨 {len(both)}행"
          f" | {a_name}만 {len(only_a)} | {b_name}만 {len(only_b)}")

    warn = self_coded(pid, [pa, pb])
    if warn:
        print(f"\n  [경고] {pid} 참여자 본인이 코딩한 파일이 있다:")
        for p in warn:
            print(f"    {p.relative_to(ROOT).as_posix()}  (마지막 저장자 = 본인)")
        print("    본인은 자기 발화 의도를 알고 있어 RST·RFL·QST 판정이 독립적이지 않다.")
        print("    아래 수치는 참고용이며 신뢰도 근거로 쓸 수 없다.")

    if not both:
        print("\n  양쪽 다 라벨이 있는 행이 없어 계산할 수 없다.")
        return

    pairs = [(A[k]["esconv"], B[k]["esconv"]) for k in both]
    po, pe, k = kappa(pairs)

    print(f"\n  단순 일치율 Po = {po:.4f}  ({sum(1 for a,b in pairs if a==b)}/{len(pairs)})")
    print(f"  우연 일치   Pe = {pe:.4f}")
    print(f"  Cohen's kappa = {k:.4f}   -> {band(k)}")

    ca, cb = Counter(a for a, _ in pairs), Counter(b for _, b in pairs)
    print(f"\n  코드별 분포와 코드별 일치")
    print(f"    {'코드':<6}{a_name:>8}{b_name:>8}{'일치':>7}{'해당코드 kappa':>16}")
    for c in CODES:
        if not (ca[c] or cb[c]):
            continue
        # 그 코드 하나만 양성으로 보는 2x2 kappa. 어느 코드가 문제인지 짚어 준다.
        bin_pairs = [(a == c, b == c) for a, b in pairs]
        _, _, bk = kappa(bin_pairs)
        agree = sum(1 for a, b in pairs if a == b == c)
        print(f"    {c:<6}{ca[c]:>8}{cb[c]:>8}{agree:>7}{bk:>16.3f}")

    print(f"\n  주요 불일치 (많은 순)")
    conf = Counter((a, b) for a, b in pairs if a != b)
    for (x, y), n in conf.most_common(8):
        print(f"    {a_name} {x:<4} <-> {b_name} {y:<4}  {n:>3}건")

    print(f"\n  불일치 문장 예시")
    shown = 0
    for key in both:
        x, y = A[key]["esconv"], B[key]["esconv"]
        if x == y:
            continue
        t = A[key]["text"]
        print(f"    turn {key[0]:>3} #{key[1]:<3} {x:<4}/{y:<4} {t[:52]}")
        shown += 1
        if shown >= 10:
            break


if __name__ == "__main__":
    main()
