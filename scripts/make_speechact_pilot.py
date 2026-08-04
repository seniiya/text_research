"""제4부 화행·감정표면성 파일럿 코딩 시트를 만든다.

사용자 발화(role=Prompt)만 모아 문장 단위로 나누고, 코더가 채울 빈 칸을 붙인다.
**라벨은 부여하지 않는다.** 이 스크립트가 하는 일은 추출·분리·표본추출뿐이다.

문장 분리는 제1부 3항(split_sentences.py)을 그대로 쓰되 두 가지를 바꾼다.

  1. 선두 불릿을 제거하지 않는다. 3항은 AI 응답의 마크다운 목록을 정리하려고
     불릿을 떼지만, 사용자 발화에서 그것은 원문 훼손이다.
  2. 목록 번호(`1.` `2.`)의 마침표를 문장 경계로 세지 않는다. 사용자가 계획을
     "…세웠던 계획은... 1. JLPT N2 응시, 2. 한능검…" 처럼 한 줄에 늘어놓는 일이
     잦은데, 그대로 두면 `1.` 하나가 문장 하나가 되어 FRAG 가 인위적으로 불어난다.

사용:
    python scripts/make_speechact_pilot.py --pid=P04
    python scripts/make_speechact_pilot.py --all            # 참여자 층화 30개
    python scripts/make_speechact_pilot.py --pid=P20 -n=30
"""

import csv
import json
import random
import re
import statistics
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import split_sentences as ss  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
CHATS = ROOT / "data" / "deid" / "chats"

# 코더가 실제로 여는 곳은 data/coding/<코더>/ 다. 제1부 시트와 같은 자리에 두되
# 파일명으로 축을 구분한다 (P04.xlsx = 제1부, P04_speechact.xlsx = 제4부).
CODING = ROOT / "data" / "coding"
# 두 코더 결과를 합쳐 κ 를 낼 때 쓰는 원본. 코더에게 주지 않는다.
MASTER_DIR = ROOT / "data" / "output" / "pilot_speechact"
CODERS = ["coder1", "coder2"]

SEED = 42
N_SAMPLE = 30

# prev_assistant_turn 은 자르지 않는다. RESP 판정은 AI 발화의 **끝**(대개 질문)을
# 봐야 갈리는데, 앞에서 잘라내면 정확히 그 부분이 사라진다. 대신 XLSX 에서 이 열만
# 줄바꿈을 꺼서 화면에는 한 줄로 보이게 하고, 셀을 클릭하면 수식 입력줄에 전문이
# 뜨도록 한다. 행 높이도 늘어나지 않는다.
CTX_LIMIT = None

# coder1 이 P01 참여자 본인이다. 자기 발화를 자기가 코딩하면 독립 판정이 아니다.
EXCLUDE = {"P01"}

# 병합용 원본의 열. note 는 코더마다 다른 것을 적으므로 한 칸에 합칠 수 없다.
# 코더 시트에서는 자기 것 하나만 보이고 이름도 그냥 note 다.
FIELDS = [
    "utterance_id", "participant_id", "turn_index", "sentence_index",
    "prev_assistant_turn", "text",
    "speech_act_coder1", "speech_act_coder2",
    "emotion_surface_coder1", "emotion_surface_coder2",
    "note_coder1", "note_coder2",
]

# 코딩북 제4부 v2. 목록 순서가 곧 판정 우선순위다.
SPEECH_ACT = [
    ("FRAG", "Fragment", "절이 완결되지 않은 발화. 길이와 무관"),
    ("RESP", "Response", "직전 AI 발화에 대한 대답·수락·거절"),
    ("PHATIC", "Phatic", "인사·작별·감사·사과"),
    ("EXPR", "Expressive", "감정·평가의 분출이 발화의 목적"),
    ("Q", "Question", "실제로 답을 요구"),
    ("C", "Command", "행동·수행을 요구"),
    ("RQ", "Rhetorical Question", "의문문 형식이나 답을 요구하지 않음"),
    ("RC", "Rhetorical Command", "명령문 형식이나 실제 지시가 아님"),
    ("S", "Statement", "정보·상태 서술"),
    ("UNC", "Unclassifiable", "위 어디에도 안 맞음. 이유를 note 에 적는다"),
]

# RESP 가 EXPR 보다 앞선다. 정서는 축 B 가 이미 잡으므로, 축 A 까지 정서를
# 기준으로 두면 두 축이 같은 것을 재게 된다.
PRIORITY = [
    "절이 완결되지 않았다 → FRAG",
    "직전 AI 발화에 대한 대답·수락·거절이다 → RESP",
    "인사·작별·감사·사과다 → PHATIC",
    "감정·평가의 분출이 발화의 목적이다 → EXPR",
    "실제로 답을 요구한다 → Q",
    "행동·수행을 요구한다 → C",
    "의문문 형식이나 답 요구가 없다 → RQ",
    "명령문 형식이나 실제 지시가 아니다 → RC",
    "정보·상태를 서술한다 → S",
    "위에 해당 없음 → UNC",
]

EMOTION = [
    ("EXP", "감정어·감정 표현이 표면에 명시됨"),
    ("IMP", "표면에 감정어는 없으나 맥락상 정서 상태가 읽힘. 단서를 note 에 적는다"),
    ("NONE", "정서와 무관한 발화"),
]

# 목록 번호. 앞이 문두·공백·쉼표류이고 뒤가 공백·줄끝이면 문장 경계가 아니다.
ORDINAL_RE = re.compile(r"(?:^|(?<=[\s,、·:：]))(\d{1,2})(\.)(?=\s|$)")

# 아래는 '분리했어야 했나' 를 사람이 보게 하는 후보 탐지용이다. 자동 분리에
# 쓰지 않는다. 한국어 종결어미는 문중 형태와 겹쳐서(`다`=부사, `지`=어미/명사)
# 정규식으로 자르면 원문이 깨진다. 판단은 코더가 한다.
FINAL_ENDING_RE = re.compile(
    r"(?:습니다|합니다|이에요|예요|거든|더라고|더라구|잖아|자나|는데|하자|"
    r"[가-힣](?:다|요|까|네|지|어|아|야|군|죠|음|슴))\s+(?=[가-힣])")


# ───────────────────────── 문장 분리 (사용자 발화용) ─────────────────────────

def mask_user_line(line):
    """3항 마스킹 + 목록 번호 마스킹. 길이는 보존한다."""
    masked = list(ss.mask_non_boundaries(line))
    for m in ORDINAL_RE.finditer(line):
        masked[m.start(2)] = ss.MASK
    return "".join(masked)


def boundaries(masked):
    ends, i = [], 0
    while i < len(masked):
        if masked[i] in ss.SENT_END:
            j = i
            while j + 1 < len(masked) and masked[j + 1] in ss.SENT_END:
                j += 1
            ends.append(j)
            i = j + 1
        else:
            i += 1
    return ends


def split_user_line(line):
    """3항 규칙 2 — 경계가 2개 이상일 때만 추가로 나눈다."""
    ends = boundaries(mask_user_line(line))
    if len(ends) < 2:
        return [line.strip()] if line.strip() else []

    out, prev = [], 0
    for e in ends:
        chunk = line[prev:e + 1].strip()
        if chunk:
            out.append(chunk)
        prev = e + 1
    tail = line[prev:].strip()
    if tail:
        out.append(tail)
    return out


def split_user_turn(text):
    """사용자 발화 1건 -> 문장 리스트. 불릿을 떼지 않는다 (원문 유지)."""
    out = []
    for raw in text.split("\n"):
        line = raw.strip()
        if not line:
            continue
        out.extend(split_user_line(line))
    return out


# ─────────────────────────────── 수집 ───────────────────────────────

def flatten(text):
    return " ".join(l.strip() for l in text.split("\n") if l.strip())


def collect(pid):
    """한 참여자의 사용자 문장을 전부 뽑는다."""
    doc = json.loads((CHATS / f"{pid}.json").read_text(encoding="utf-8"))
    rows, turn, prev_assistant = [], 0, ""

    for msg in doc["messages"]:
        if msg["role"] == "Response":
            prev_assistant = flatten(msg["say"])
            continue
        if msg["role"] != "Prompt":
            continue

        turn += 1
        ctx = prev_assistant
        if CTX_LIMIT and len(ctx) > CTX_LIMIT:
            ctx = ctx[:CTX_LIMIT] + "…"

        for n, sent in enumerate(split_user_turn(msg["say"]), start=1):
            rows.append({
                "utterance_id": f"{pid}-t{turn:02d}-s{n:02d}",
                "participant_id": pid,
                "turn_index": turn,
                "sentence_index": n,
                # 맥락은 턴 단위다. 한 턴이 여러 문장이면 첫 문장에만 적는다.
                # 같은 응답을 다섯 번 읽게 하지 않기 위함이다. 대신 행 순서를
                # 바꾸면 뒤 문장이 맥락을 잃으므로 시트를 정렬하지 않는다.
                "prev_assistant_turn": ctx if n == 1 else "",
                "text": sent,
                "speech_act_coder1": "",
                "speech_act_coder2": "",
                "emotion_surface_coder1": "",
                "emotion_surface_coder2": "",
                "note_coder1": "",
                "note_coder2": "",
            })
    return rows


def collect_all():
    pool = {}
    for f in sorted(CHATS.glob("P*.json")):
        if f.stem in EXCLUDE:
            continue
        rows = collect(f.stem)
        if rows:
            pool[f.stem] = rows
    return pool


# ─────────────────────────────── 표본추출 ───────────────────────────────

def sample_one(rows, n, rng):
    if len(rows) <= n:
        return list(rows), True          # 전수
    return rng.sample(rows, n), False


def sample_stratified(pool, n, rng):
    """참여자별 최소 1개를 먼저 채우고 나머지를 무작위로 채운다."""
    picked, rest = [], []
    for pid in sorted(pool):
        rows = list(pool[pid])
        rng.shuffle(rows)
        picked.append(rows[0])
        rest.extend(rows[1:])
    if len(picked) >= n:
        return rng.sample(picked, n), False
    picked.extend(rng.sample(rest, n - len(picked)))
    return picked, False


def order(rows):
    return sorted(rows, key=lambda r: (r["participant_id"],
                                       r["turn_index"], r["sentence_index"]))


# ─────────────────────────────── 출력 ───────────────────────────────

def write_csv(rows, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    return path


def coder_cols(coder):
    return ["utterance_id", "participant_id", "turn_index", "sentence_index",
            "prev_assistant_turn", "text",
            f"speech_act_{coder}", f"emotion_surface_{coder}", "note"]


def coder_view(rows, coder):
    """코더 시트용 행. 그 코더의 note_<coder> 를 note 라는 이름으로 보여 준다."""
    return [{**r, "note": r.get(f"note_{coder}", "")} for r in rows]


def carry_over(base, coder, rows):
    """이미 입력한 라벨과 note 를 새 시트로 옮긴다.

    문장 분리 규칙이나 표본이 바뀌면 행이 새로 만들어진다. 그때 입력해 둔 것이
    사라지지 않도록 기존 XLSX 를 읽어 되돌려 놓는다. (utterance_id, text) 가
    둘 다 같아야 옮긴다. 문장이 달라진 행은 옮기지 않는다 — 잘못된 위치에
    라벨이 붙는 것보다 비는 편이 낫다 (제1부 carry_over 와 같은 원칙).

    코더가 엑셀에서 작업하므로 XLSX 를 읽는다. CSV 는 sync_labels.py 를
    돌리기 전까지 뒤처져 있다.
    """
    path = CODING / coder / f"{base}.xlsx"
    if not path.exists():
        return 0, 0
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return 0, 0

    ws = load_workbook(path)["coding"]
    head = [c.value for c in ws[1]]
    if "utterance_id" not in head or "text" not in head:
        return 0, 0
    at = {h: i for i, h in enumerate(head)}

    # 엑셀 열 이름 -> 행 딕셔너리 키. note 는 코더별 키로 바꿔 담는다.
    pairs = [(f, f) for f in (f"speech_act_{coder}", f"emotion_surface_{coder}")
             if f in at]
    if "note" in at:
        pairs.append(("note", f"note_{coder}"))

    old = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        vals = {key: str(r[at[col]]).strip()
                for col, key in pairs if r[at[col]] not in (None, "")}
        if vals:
            old[(r[at["utterance_id"]], r[at["text"]])] = vals

    kept = 0
    for row in rows:
        vals = old.pop((row["utterance_id"], row["text"]), None)
        if vals:
            row.update(vals)
            kept += 1
    return kept, len(old)


def write_coder_csv(rows, path, coder):
    """제1부 시트와 같이 CSV·XLSX 를 나란히 둔다. 라벨 정본은 XLSX 다."""
    cols = coder_cols(coder)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return path


def write_xlsx(rows, path, coder):
    """코더 1인용 시트. 자기 열만 들어간다 (코딩북 제2부 15항 — 블라인드)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError:
        return None

    cols = coder_cols(coder)
    label_cols = [f"speech_act_{coder}", f"emotion_surface_{coder}", "note"]
    widths = {"utterance_id": 16, "participant_id": 8, "turn_index": 7,
              "sentence_index": 7, "prev_assistant_turn": 52, "text": 62,
              f"speech_act_{coder}": 12, f"emotion_surface_{coder}": 14, "note": 34}

    wb = Workbook()
    sheet = wb.active
    sheet.title = "coding"

    head_fill = PatternFill("solid", fgColor="2F4858")
    ctx_fill = PatternFill("solid", fgColor="F2EFE9")
    input_fill = PatternFill("solid", fgColor="FFF9E6")

    sheet.append(cols)
    for c in sheet[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 24

    for r in rows:
        sheet.append([r.get(c, "") for c in cols])

    for i, c in enumerate(cols, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = widths[c]

    # 맥락 열은 전문이 들어가므로 줄바꿈을 끈다. 화면에는 한 줄로 잘려 보이고
    # 셀을 클릭하면 수식 입력줄에 전체가 뜬다.
    wrap = {cols.index("text") + 1}
    ctx_col = cols.index("prev_assistant_turn") + 1
    idx = {c: cols.index(c) + 1 for c in label_cols}

    for i in range(2, len(rows) + 2):
        for j in range(1, len(cols) + 1):
            cell = sheet.cell(row=i, column=j)
            cell.alignment = Alignment(wrap_text=(j in wrap), vertical="top",
                                       horizontal="left" if j in wrap else "center")
            cell.font = Font(size=10, italic=(j == ctx_col),
                             color="6B6255" if j == ctx_col else "000000")
            if j == ctx_col:
                cell.fill = ctx_fill
            cell.protection = Protection(locked=True)
        for col in idx.values():
            cell = sheet.cell(row=i, column=col)
            cell.protection = Protection(locked=False)
            cell.fill = input_fill

    last = len(rows) + 1
    for field, codes, title in (
        (f"speech_act_{coder}", [c[0] for c in SPEECH_ACT], "화행 6+1"),
        (f"emotion_surface_{coder}", [c[0] for c in EMOTION], "감정 표면성"),
    ):
        dv = DataValidation(type="list", formula1='"' + ",".join(codes) + '"',
                            allow_blank=True, showDropDown=False,
                            errorTitle="코드 아님", error=" / ".join(codes),
                            promptTitle=title, prompt=" ".join(codes))
        sheet.add_data_validation(dv)
        letter = get_column_letter(idx[field])
        dv.add(f"{letter}2:{letter}{last}")

    sheet.freeze_panes = "A2"
    sheet.protection.sheet = True
    sheet.protection.formatColumns = False
    sheet.protection.formatRows = False
    # 제1부 시트와 달리 정렬·필터를 잠근다. 맥락이 턴의 첫 문장에만 있어서
    # 행 순서가 바뀌면 뒤 문장이 무슨 응답에 대한 반응인지 알 수 없게 된다.
    sheet.protection.sort = True
    sheet.protection.autoFilter = True

    ref = wb.create_sheet("코드표")
    ref.append(["축 A. 화행 (Speech Act) — 9클래스 + UNC"])
    ref.append(["코드", "이름", "기준"])
    for row in SPEECH_ACT:
        ref.append(list(row))
    ref.append([])
    ref.append(["판정 우선순위 — 위에서부터 먼저 걸리는 규칙을 따른다"])
    for n, line in enumerate(PRIORITY, start=1):
        ref.append([n, line])
    ref.append([])
    ref.append(["축 B. 감정 표면성 (Emotion Surface)"])
    ref.append(["코드", "기준"])
    for row in EMOTION:
        ref.append(list(row))
    ref.append([])
    ref.append(["판정은 문장 형식이 아니라 실제 기능으로 한다."])
    ref.append(["\"모닝콜 해줄 수 있어?\" -> 형식은 의문문, 기능은 요청 -> C"])
    ref.append(["\"내가 뭘 그렇게 잘못했냐\" -> 형식은 의문문, 답 요구 없음 -> RQ"])
    ref.append(["\"소고기\" (AI 가 \"뭐 먹어?\" 라고 물은 직후) -> 규칙 2 -> RESP"])
    ref.append([])
    ref.append(["FRAG 와 UNC 는 막히는 지점이 다르다."])
    ref.append(["  FRAG = 형태. 절이 완결되지 않음. 길이와 무관"])
    ref.append(["  UNC  = 기능. 완결됐는데 어디에도 안 들어감"])
    ref.append(["UNC 를 억지로 채우지 않는다. UNC 비율 자체가 결과다."])
    ref.append(["IMP 로 판정하면 note 에 단서를 한 문장으로 적는다."])
    for i, w in enumerate([10, 22, 60], start=1):
        ref.column_dimensions[get_column_letter(i)].width = w
    for row in ref.iter_rows():
        for c in row:
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ─────────────────────────────── 리포트 ───────────────────────────────

def section(title):
    print()
    print("=" * 74)
    print(title)
    print("=" * 74)


def report(pool, sample, census, target):
    total = sum(len(v) for v in pool.values())
    section("1. 모집단")
    print(f"대상 참여자 : {len(pool)}명  (제외: {', '.join(sorted(EXCLUDE))})")
    print(f"user 턴     : {sum(r['sentence_index'] == 1 for v in pool.values() for r in v)}")
    print(f"user 문장   : {total}")

    section("2. 표본")
    print(f"방식 : {'전수 (문장 수가 목표보다 적음)' if census else f'무작위 seed={SEED}'}")
    print(f"크기 : {len(sample)} / 목표 {target}")
    print()
    dist = {}
    for r in sample:
        dist[r["participant_id"]] = dist.get(r["participant_id"], 0) + 1
    print(f"{'참여자':8}{'문장':>6}{'모집단':>8}{'비율':>8}")
    print("-" * 30)
    for pid in sorted(dist):
        n, m = dist[pid], len(pool[pid])
        print(f"{pid:8}{n:>6}{m:>8}{n / m * 100:>7.0f}%")

    lens = [len(r["text"]) for r in sample]
    section("3. 문장 길이 (글자 수, 공백 포함)")
    print(f"평균 : {statistics.mean(lens):.1f}")
    print(f"중위 : {statistics.median(lens):.0f}")
    print(f"최소 : {min(lens)}   최대 : {max(lens)}")
    print(f"모집단 평균 : "
          f"{statistics.mean([len(r['text']) for v in pool.values() for r in v]):.1f}")

    section("4. 검토 필요 — 종결어미로 더 쪼갤 수 있는 문장 (자동 분리 안 함)")
    hits = [r for r in sample if FINAL_ENDING_RE.search(r["text"])]
    for r in hits:
        print(f"  {r['utterance_id']}  {r['text'][:64]}")
    print(f"\n소계: {len(hits)}건 / {len(sample)}")
    print("한국어 종결어미는 문중 형태와 겹쳐 자동 분리하면 원문이 깨진다.")
    print("여기 걸린 문장을 더 나눌지는 코더가 조정 회의에서 정한다.")

    section("5. 가명화 자리표시자가 남은 문장")
    ph = [r for r in sample if re.search(r"\[(이름|이미지|링크|Attachment[^\]]*)\]", r["text"])]
    for r in ph:
        print(f"  {r['utterance_id']}  {r['text'][:64]}")
    print(f"\n소계: {len(ph)}건")


def main():
    args = sys.argv[1:]
    flags = {a.split("=")[0].lstrip("-"): a.split("=", 1)[-1]
             for a in args if a.startswith("-")}

    target = int(flags.get("n", N_SAMPLE))
    rng = random.Random(SEED)
    pool = collect_all()

    if "all" in flags:
        pid = None
        sample, census = sample_stratified(pool, target, rng)
    else:
        pid = flags.get("pid", "P04")
        if pid in EXCLUDE:
            sys.exit(f"{pid} 은(는) 제외 대상이다 ({', '.join(sorted(EXCLUDE))}).")
        if pid not in pool:
            sys.exit(f"{pid} 로그가 없다.")
        sample, census = sample_one(pool[pid], target, rng)

    sample = order(sample)
    view = pool if pid is None else {pid: pool[pid]}
    report(view, sample, census, target)

    base = f"{pid}_speechact" if pid else "pilot_speechact"

    section("6. 산출물 — 코딩용 (코더가 여는 파일)")
    for coder in CODERS:
        out = CODING / coder
        kept, lost = carry_over(base, coder, sample)
        view = coder_view(sample, coder)
        c = write_coder_csv(view, out / f"{base}.csv", coder)
        x = write_xlsx(view, out / f"{base}.xlsx", coder)
        print(f"[{coder}]")
        if kept or lost:
            print(f"  기존 입력 {kept}행 이어받음"
                  + (f" / {lost}행은 문장이 바뀌어 옮기지 못함 ⚠" if lost else ""))
        print(f"  XLSX : {x.relative_to(ROOT).as_posix()}   <- 여기에 입력한다 (드롭다운)"
              if x else f"  XLSX : 건너뜀 — openpyxl 없음")
        print(f"  CSV  : {c.relative_to(ROOT).as_posix()}   (Git 확인용 사본)")

    # 병합용 원본은 --master 를 줄 때만 만든다. 두 코더 열이 한 파일에 있어
    # κ 산출 때만 쓰는데, coder2 코딩 전에는 빈 파일이 참여자 수만큼 쌓인다.
    if "master" in flags:
        master = write_csv(sample, MASTER_DIR / f"pilot_coding_sheet_{pid or 'all'}.csv")
        section("7. 산출물 — 병합용 (코더에게 주지 않는다)")
        print(f"CSV  : {master.relative_to(ROOT).as_posix()}")
        print("두 코더 열이 한 파일에 있다. κ 산출 때만 쓴다 (코딩북 제2부 15항).")
    print()
    print("이 스크립트는 라벨을 부여하지 않는다. 채워져 있다면 전부 코더가 넣은 것이다.")


if __name__ == "__main__":
    main()
