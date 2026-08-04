"""제4부 코딩북 v1(3i4K 6+UNC) 과 v2(9+UNC) 라벨을 나란히 놓고 전이를 센다.

왜 필요한가 — v2 에서 새로 넣은 PHATIC·RESP·EXPR 은 원 체계(3i4K)에 없던 코드다.
원 체계를 그대로 적용하면 그 발화들이 어디로 갔는지를 보여야, "체계가 안 맞아서
고쳤다"가 인상이 아니라 근거가 된다. 그 대조표를 만든다.

v1 라벨이 남아 있는 참여자만 대상이다. v2 로만 코딩한 참여자는 원 체계 값을
복원할 수 없다 (RESP·PHATIC·EXPR 은 S·FRAG·Q·C 어디로든 갈 수 있다).

    data/coding/<코더>/P##_speechact_v1.xlsx   v1 라벨 (보존본)
    data/coding/<코더>/P##_speechact.xlsx      v2 라벨 (현행)
    -> data/output/pilot_speechact/P##_v1_v2_compare.csv

사용: python scripts/compare_speechact_versions.py P04 [--coder=coder1]
"""

import csv
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CODING = ROOT / "data" / "coding"
OUT = ROOT / "data" / "output" / "pilot_speechact"

V1_CODES = ["FRAG", "S", "Q", "C", "RQ", "RC", "UNC"]
V2_CODES = ["FRAG", "RESP", "PHATIC", "EXPR", "Q", "C", "RQ", "RC", "S", "UNC"]
ADDED = {"RESP", "PHATIC", "EXPR"}      # v2 에서 새로 넣은 코드

FIELDS = ["utterance_id", "turn_index", "sentence_index", "text",
          "speech_act_v1", "speech_act_v2", "changed",
          "emotion_surface", "note"]


def read_sheet(path, coder):
    from openpyxl import load_workbook

    ws = load_workbook(path)["coding"]
    head = [c.value for c in ws[1]]
    at = {h: i for i, h in enumerate(head)}
    need = ("utterance_id", "text", f"speech_act_{coder}")
    missing = [c for c in need if c not in at]
    if missing:
        sys.exit(f"[중단] {path.name} 에 열이 없다: {', '.join(missing)}")

    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        uid = r[at["utterance_id"]]
        if not uid:
            continue
        get = lambda c: (str(r[at[c]]).strip() if c in at and r[at[c]] else "")  # noqa: E731
        out[uid] = {
            "text": r[at["text"]] or "",
            "act": get(f"speech_act_{coder}"),
            "emotion": get(f"emotion_surface_{coder}"),
            "note": get("note"),
        }
    return out


def build(pid, coder):
    v1_path = CODING / coder / f"{pid}_speechact_v1.xlsx"
    v2_path = CODING / coder / f"{pid}_speechact.xlsx"
    if not v1_path.exists():
        sys.exit(f"[중단] v1 보존본이 없다: {v1_path.relative_to(ROOT).as_posix()}\n"
                 f"       v1 으로 코딩한 참여자만 대조할 수 있다.")
    if not v2_path.exists():
        sys.exit(f"[중단] v2 시트가 없다: {v2_path.relative_to(ROOT).as_posix()}")

    v1, v2 = read_sheet(v1_path, coder), read_sheet(v2_path, coder)

    rows, skipped = [], 0
    for uid in sorted(set(v1) & set(v2)):
        a, b = v1[uid], v2[uid]
        if a["text"] != b["text"]:
            skipped += 1        # 문장이 달라졌으면 짝이 아니다
            continue
        parts = uid.rsplit("-t", 1)[-1].split("-s")
        rows.append({
            "utterance_id": uid,
            "turn_index": parts[0],
            "sentence_index": parts[1] if len(parts) > 1 else "",
            "text": b["text"],
            "speech_act_v1": a["act"],
            "speech_act_v2": b["act"],
            "changed": "Y" if a["act"] != b["act"] else "",
            "emotion_surface": b["emotion"],
            "note": b["note"],
        })
    return rows, skipped


def report(pid, rows, skipped):
    paired = [r for r in rows if r["speech_act_v1"] and r["speech_act_v2"]]
    print(f"\n{'='*70}\n{pid} — v1(3i4K 6+UNC) vs v2(9+UNC)\n{'='*70}")
    print(f"대조 가능 {len(paired)}문장 / 전체 {len(rows)}"
          + (f"  (문장이 달라 제외 {skipped})" if skipped else ""))
    if not paired:
        return

    changed = [r for r in paired if r["changed"]]
    print(f"코드가 바뀐 문장 {len(changed)} ({len(changed)/len(paired)*100:.0f}%)")

    print("\n■ 전이  v1 -> v2")
    trans = {}
    for r in paired:
        trans.setdefault(r["speech_act_v1"], Counter())[r["speech_act_v2"]] += 1
    for a in [c for c in V1_CODES if c in trans]:
        moved = ", ".join(f"{b} {n}" for b, n in trans[a].most_common())
        print(f"  {a:5} {sum(trans[a].values()):>3}  ->  {moved}")

    print("\n■ v2 에서 새로 넣은 코드가 원 체계에서 어디에 있었나")
    for code in [c for c in V2_CODES if c in ADDED]:
        src = Counter(r["speech_act_v1"] for r in paired if r["speech_act_v2"] == code)
        if src:
            print(f"  {code:7} {sum(src.values()):>3}  <-  "
                  + ", ".join(f"{k} {v}" for k, v in src.most_common()))

    print("\n■ 분포")
    c1 = Counter(r["speech_act_v1"] for r in paired)
    c2 = Counter(r["speech_act_v2"] for r in paired)
    n = len(paired)
    print(f"  {'코드':8}{'v1':>8}{'v2':>8}")
    for code in V2_CODES:
        if c1[code] or c2[code]:
            a = f"{c1[code]} ({c1[code]/n*100:.0f}%)" if c1[code] else "-"
            b = f"{c2[code]} ({c2[code]/n*100:.0f}%)" if c2[code] else "-"
            print(f"  {code:8}{a:>8}{b:>8}")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    flags = {a.split("=")[0].lstrip("-"): a.split("=", 1)[-1]
             for a in sys.argv[1:] if a.startswith("-")}
    pid = args[0] if args else "P04"
    coder = flags.get("coder", "coder1")

    rows, skipped = build(pid, coder)
    report(pid, rows, skipped)

    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / f"{pid}_v1_v2_compare.csv"
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    print(f"\n대조표: {path.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
